# cording = utf-8
import yaml
import openpyxl
import json

from collections import OrderedDict


def read_yaml(yaml_file):
    with open(yaml_file) as yml:
        data = yaml.load(yml, Loader=yaml.FullLoader)
    return data


def yaml2dict(yaml):
    pass


def pyxl2json(pyxl_file):
    wb = openpyxl.load_workbook(pyxl_file)
    ws = wb.active

    col_set = []
    for col in ws[1]:
        if col.value is None:
            pass
        else:
            col_set.append(col.value)
    data = OrderedDict()
    data_dict = OrderedDict()

    for row in range(2, ws.max_row + 1):
        for cell in range(2, len(col_set) + 1):
            data[col_set[cell-1]] = ws.cell(row=row, column=cell).value
        data_dict[ws.cell(row=row, column=1).value] = data
        data = OrderedDict()

    j1 = json.dumps(data_dict, ensure_ascii=False)
    j2 = json.loads(j1)

    return j2